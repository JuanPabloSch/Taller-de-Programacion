# views.py (fusionado y limpio)
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_http_methods
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.contrib import messages
from django.db import transaction
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime, date, timedelta
from dateutil.relativedelta import relativedelta
from decimal import Decimal, ROUND_HALF_UP
import csv, os, openpyxl
from openpyxl.utils import get_column_letter
from fpdf import FPDF
from django.utils import timezone
from .models import (
    PlanPago,
    Cuota,
    Regularizacion,
    ReglaEstructura,
    ReglaMora,
    CuotaRegularizacion,
    HistorialAccion
)
from .forms import (
    PlanPagoForm,
    RegularizacionForm,
    ReglaEstructuraForm,
    ReglaEstructuraRegularizacionForm,
    ReglaMoraForm
)
from .decorators import group_required, can_delete, can_modify

def registrar_accion(usuario, accion, plan=None, regularizacion=None, descripcion=""):
    # Registrar la acción en el historial
    # El campo 'plan' puede ser None para acciones generales
    HistorialAccion.objects.create(
        usuario=usuario,
        accion=accion,
        plan=plan,  # Puede ser None
        descripcion=descripcion
    )

# -------------------------------
# Clase PDF personalizada
# -------------------------------
class PDF(FPDF):
    def header(self):
        logo_path = os.path.join("static", "img", "logo.png")
        if os.path.exists(logo_path):
            try:
                self.image(logo_path, 10, 8, 20)  # (x, y, ancho)
            except Exception:
                # si hay problemas con la imagen, no rompe el PDF
                pass
        self.set_font("Arial", "B", 14)
        self.cell(0, 10, "ISDM - Sistema de Pagos", border=False, ln=1, align="C")
        self.ln(5)

    def footer(self):
        self.set_y(-15)
        self.set_font("Arial", "I", 8)
        self.cell(0, 10, f"Página {self.page_no()}/{{nb}}", align="C")


# -------------------------------
# Página de inicio
# -------------------------------
@login_required
def home(request):
    return render(request, "index.html")


# -------------------------------
# CRUD Planes (lista principal)
# -------------------------------
@login_required
def planes_list(request):
    # Instanciar los formularios para Planes Normales con prefijos únicos
    regularizacion_form = RegularizacionForm(prefix='regularizacion')
    estructura_form = ReglaEstructuraForm(prefix='estructura')
    mora_form = ReglaMoraForm(prefix='mora')

    # Instanciar formularios para Regularizaciones con prefijos diferentes
    regularizacion_reg_form = RegularizacionForm(prefix='regularizacion_reg')
    estructura_reg_form = ReglaEstructuraRegularizacionForm(prefix='estructura_reg')
    mora_reg_form = ReglaMoraForm(prefix='mora_reg')

    context = {
        # Formularios para Planes Normales
        'regularizacion_form': regularizacion_form,
        'estructura_form': estructura_form,
        'mora_form': mora_form,
        # Formularios para Regularizaciones
        'regularizacion_reg_form': regularizacion_reg_form,
        'estructura_reg_form': estructura_reg_form,
        'mora_reg_form': mora_reg_form,
    }

    return render(request, "planes_list.html", context)


# views.py

# ... (código previo)

@login_required
def planes_data(request):
    """
    Endpoint JSON para datatables / fetch de planes activos.
    Combina PlanPago y Regularizacion activos.
    """
    
    # 1. Obtener Planes de Pago activos
    planes = PlanPago.objects.filter(estado='A').values(
        'id', 'tipo', 'nombre', 'carrera', 'cohorte', 'modalidad'
    )
    
    # 2. Obtener Regularizaciones activas
    # NOTA: Asumimos que Regularizacion tiene 'nombre', 'carrera', 'cohorte', 'modalidad'.
    # Si no los tiene, usa .annotate() para mapear campos.
    regularizaciones = Regularizacion.objects.filter(estado='A').values(
        'id', 'tipo', 'nombre', 'carrera', 'cohorte', 'modalidad'
    )
    
    # 3. Combinar y formatear
    combined_data = list(planes) + list(regularizaciones)

    data = [
        {
            "id": p['id'],
            # Si el modelo tiene un campo 'tipo', úsalo. Si no, forzamos un valor por defecto.
            "tipo": p.get('tipo', 'plan') if 'cohorte' in p else p.get('tipo', 'regularizacion'), 
            "nombre": p['nombre'],
            "carrera": p['carrera'],
            "cohorte": p['cohorte'],
            "modalidad": p['modalidad'],
        }
        for p in combined_data
    ]
    return JsonResponse({"data": data})

# ... (código posterior)

@require_POST
@login_required
def plan_guardar(request):
    """
    Crear o actualizar plan vía POST (AJAX u ordinario).
    """
    if not can_modify(request.user):
        return JsonResponse({"ok": False, "msg": "No tienes permisos para realizar esta acción"}, status=403)

    datos = request.POST
    plan_id = datos.get("id")

    if plan_id:
        try:
            plan = PlanPago.objects.get(pk=plan_id)
            plan.nombre = datos.get("nombre", plan.nombre)
            plan.carrera = datos.get("carrera", plan.carrera)
            plan.cohorte = datos.get("cohorte", plan.cohorte)
            plan.modalidad = datos.get("modalidad", plan.modalidad)
            plan.save()
            return JsonResponse({"ok": True, "msg": "Plan actualizado"})
        except PlanPago.DoesNotExist:
            return JsonResponse({"ok": False, "msg": "Plan no encontrado"}, status=404)
    else:
        PlanPago.objects.create(
            nombre=datos.get("nombre", ""),
            carrera=datos.get("carrera", ""),
            cohorte=datos.get("cohorte", ""),
            modalidad=datos.get("modalidad", ""),
            estado='S'  # por defecto guardamos como suspendido para revisión del admin
        )
        return JsonResponse({"ok": True, "msg": "Plan creado"})


@require_POST
@login_required
def plan_borrar(request, pk):
    """
    Marcar plan como inactivo (iEstado=False) - borrado lógico.
    """
    if not can_delete(request.user):
        return JsonResponse({"ok": False, "msg": "No tienes permisos para eliminar planes"}, status=403)

    try:
        plan = PlanPago.objects.get(pk=pk)
        plan.iEstado = False
        plan.save()
        return JsonResponse({"ok": True, "msg": "Plan eliminado"})
    except PlanPago.DoesNotExist:
        return JsonResponse({"ok": False, "msg": "Plan no encontrado"}, status=404)


# -------------------------------
# FORMULARIOS DE PLANES (modal AJAX)
# -------------------------------
@login_required
@group_required('Administrador', 'Tesorero')
@require_http_methods(["POST"])
def plan_crear(request):
    """Crear un nuevo Plan de Pago (no regularización) con reglas de estructura y mora"""
    if request.method == 'POST':
        print("=" * 50)
        print("VIEW: plan_crear (PLAN NORMAL)")
        print("POST keys:", list(request.POST.keys()))
        print("=" * 50)

        regularizacion_form = RegularizacionForm(request.POST, prefix='regularizacion')
        estructura_form = ReglaEstructuraForm(request.POST, prefix='estructura')
        mora_form = ReglaMoraForm(request.POST, prefix='mora')

        # Hacer que el campo 'tipo' no sea requerido para planes normales (lo establecemos manualmente)
        if 'tipo' in regularizacion_form.fields:
            regularizacion_form.fields['tipo'].required = False

        if all([regularizacion_form.is_valid(), estructura_form.is_valid(), mora_form.is_valid()]):
            try:
                with transaction.atomic():
                    regularizacion = regularizacion_form.save(commit=False)
                    regularizacion.tipo = 'Plan Normal'  # Importante: marcar como Plan Normal
                    regularizacion.estado = 'S'  # Se marca como suspendido
                    print(f"DEBUG: Guardando plan con tipo: {regularizacion.tipo}")  # Debug
                    regularizacion.save()
                    print(f"DEBUG: Plan guardado con ID: {regularizacion.id}, tipo: {regularizacion.tipo}")  # Debug

                    estructura = estructura_form.save(commit=False)
                    estructura.regularizacion = regularizacion
                    estructura.save()

                    mora = mora_form.save(commit=False)
                    mora.regularizacion = regularizacion
                    mora.save()

                    # Guardar las cuotas si fueron enviadas
                    cuotas_json = request.POST.get('cuotas_json')
                    if cuotas_json:
                        import json
                        from datetime import datetime
                        try:
                            cuotas_data = json.loads(cuotas_json)
                            for cuota in cuotas_data:
                                # Parsear la fecha de vencimiento
                                fecha_vto = None
                                if cuota.get('vto'):
                                    try:
                                        # Intentar formato DD/MM/YYYY
                                        fecha_vto = datetime.strptime(cuota['vto'], '%d/%m/%Y').date()
                                    except:
                                        try:
                                            # Intentar formato YYYY-MM-DD (ISO)
                                            fecha_vto = datetime.strptime(cuota['vto'], '%Y-%m-%d').date()
                                        except:
                                            pass

                                CuotaRegularizacion.objects.create(
                                    regularizacion=regularizacion,
                                    numero_cuota=cuota.get('nro', 0),
                                    fecha_vencimiento=fecha_vto,
                                    monto_base=cuota.get('base', 0),
                                    monto_interes=cuota.get('interes', 0),
                                    monto_cuota=cuota.get('monto_total', 0),
                                    monto_mora=cuota.get('mora', 0),
                                    estado='P'  # Pendiente por defecto
                                )
                        except Exception as e:
                            # Si hay error al guardar cuotas, log pero no falla la transacción
                            print(f"Error al guardar cuotas: {e}")

                    # Registrar en historial
                    registrar_accion(
                        usuario=request.user,
                        accion="Creó un plan",
                        plan=None,
                        descripcion=f"Plan Normal creado: '{regularizacion.nombre}' - {regularizacion.carrera} ({regularizacion.cohorte}) - Suspendido"
                    )

                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({'success': True, 'message': 'Plan creado correctamente.'})

                messages.success(request, 'El plan fue creado exitosamente.')
                return redirect('planes_list')

            except Exception as e:
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': str(e)}, status=500)
                messages.error(request, f'Ocurrió un error al guardar: {e}')

        else:
            # Responder errores (AJAX o normal)
            print("DEBUG: Formularios con errores")
            print(f"Errores regularizacion: {regularizacion_form.errors}")
            print(f"Errores estructura: {estructura_form.errors}")
            print(f"Errores mora: {mora_form.errors}")

            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': {
                        'regularizacion': regularizacion_form.errors,
                        'estructura': estructura_form.errors,
                        'mora': mora_form.errors
                    }
                }, status=400)

            messages.error(request, 'Hay errores en los formularios.')

    return redirect('planes_list')


# Regularizaciones (combinado con reglas)
@require_http_methods(["POST"])
def regularizacion_crear(request):
    if request.method == 'POST':
        regularizacion_form = RegularizacionForm(request.POST, prefix='regularizacion_reg')
        estructura_form = ReglaEstructuraRegularizacionForm(request.POST, prefix='estructura_reg')
        mora_form = ReglaMoraForm(request.POST, prefix='mora_reg')

        if all([regularizacion_form.is_valid(), estructura_form.is_valid(), mora_form.is_valid()]):
            try:
                with transaction.atomic():
                    regularizacion = regularizacion_form.save(commit=False)
                    regularizacion.estado = 'S'  # Se marca como suspendido
                    regularizacion.save()

                    estructura = estructura_form.save(commit=False)
                    estructura.regularizacion = regularizacion
                    estructura.save()

                    mora = mora_form.save(commit=False)
                    mora.regularizacion = regularizacion
                    mora.save()

                    # Guardar las cuotas si fueron enviadas
                    cuotas_json = request.POST.get('cuotas_json')
                    if cuotas_json:
                        import json
                        from .models import CuotaRegularizacion
                        from datetime import datetime
                        try:
                            cuotas_data = json.loads(cuotas_json)
                            for cuota in cuotas_data:
                                # Parsear la fecha de vencimiento
                                fecha_vto = None
                                if cuota.get('vto'):
                                    try:
                                        # Intentar formato DD/MM/YYYY
                                        fecha_vto = datetime.strptime(cuota['vto'], '%d/%m/%Y').date()
                                    except:
                                        try:
                                            # Intentar formato YYYY-MM-DD (ISO)
                                            fecha_vto = datetime.strptime(cuota['vto'], '%Y-%m-%d').date()
                                        except:
                                            pass

                                CuotaRegularizacion.objects.create(
                                    regularizacion=regularizacion,
                                    numero_cuota=cuota.get('nro', 0),
                                    fecha_vencimiento=fecha_vto,
                                    monto_base=cuota.get('base', 0),
                                    monto_interes=cuota.get('interes', 0),
                                    monto_cuota=cuota.get('monto_total', 0),
                                    monto_mora=cuota.get('mora', 0),
                                    estado='P'  # Pendiente por defecto
                                )
                        except Exception as e:
                            # Si hay error al guardar cuotas, log pero no falla la transacción
                            print(f"Error al guardar cuotas: {e}")

                    # Registrar en historial
                    registrar_accion(
                        usuario=request.user,
                        accion="Creó una regularización",
                        plan=None,
                        descripcion=f"Regularización creada: '{regularizacion.nombre}' - {regularizacion.carrera} ({regularizacion.cohorte}) - Suspendida"
                    )

                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({'success': True, 'message': 'Regularización creada correctamente.'})

                messages.success(request, 'La regularización fue creada exitosamente.')
                return redirect('planes_suspendidos')

            except Exception as e:
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': str(e)}, status=500)
                messages.error(request, f'Ocurrió un error al guardar: {e}')

        else:
            # Responder errores (AJAX o normal)
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': {
                        'regularizacion_reg': regularizacion_form.errors,
                        'estructura_reg': estructura_form.errors,
                        'mora_reg': mora_form.errors
                    }
                }, status=400)

            messages.error(request, 'Hay errores en los formularios.')
            # No guardamos parcialmente en este flujo; redirigimos
    return redirect('planes_list')


@login_required
def regularizacion_editar(request, pk):
    """
    Editar una regularización existente con sus reglas de estructura y mora.
    Devuelve los datos del plan en formato JSON para cargar en el modal.
    Solo Administrador y Tesorero pueden hacer POST (editar).
    Consulta puede hacer GET (solo lectura).
    """
    # Verificar permisos para POST (edición)
    if request.method == 'POST':
        if not (request.user.is_superuser or request.user.groups.filter(name__in=['Administrador', 'Tesorero']).exists()):
            return JsonResponse({'ok': False, 'msg': 'No tiene permisos para editar regularizaciones'}, status=403)

    regularizacion = get_object_or_404(Regularizacion, pk=pk)

    if request.method == 'GET':
        # Obtener las reglas asociadas
        try:
            estructura = ReglaEstructura.objects.get(regularizacion=regularizacion)
        except ReglaEstructura.DoesNotExist:
            estructura = None

        try:
            mora = ReglaMora.objects.get(regularizacion=regularizacion)
        except ReglaMora.DoesNotExist:
            mora = None

        # Obtener las cuotas existentes
        cuotas = CuotaRegularizacion.objects.filter(regularizacion=regularizacion).order_by('numero_cuota')

        # Preparar datos para enviar al frontend
        data = {
            'ok': True,
            'regularizacion': {
                'id': regularizacion.id,
                'nombre': regularizacion.nombre,
                'carrera': regularizacion.carrera,
                'modalidad': regularizacion.modalidad,
                'cohorte': regularizacion.cohorte,
            },
            'estructura': {
                'origen_deuda': estructura.origen_deuda if estructura else '',
                'valor': str(estructura.valor) if estructura else '',
                'tasa': str(estructura.tasa) if estructura else '',
                'pago_inicial': str(estructura.pago_incial) if estructura else '',
                'cantidad_cuotas': estructura.cantidad_de_cuotas if estructura else '',
                'frecuencia': estructura.frecuencia_de_pago if estructura else '',
                'dia_vencimiento': estructura.dia_vencimiento if estructura else '',
            } if estructura else None,
            'mora': {
                'tipo_recargo': mora.tipo_de_recargo if mora else '',
                'cantidad': str(mora.cantidad_recargo) if mora else '',
                'frecuencia': mora.frecuencia_aplicacion if mora else '',
                'dias_gracia': mora.dias_gracia if mora else '',
                'veces_aplicacion': mora.veces_aplicacion if mora else '',
            } if mora else None,
            'cuotas': [
                {
                    'nro': c.numero_cuota,
                    'base': str(c.monto_base),
                    'interes': str(c.monto_interes),
                    'monto_total': str(c.monto_cuota),
                    'mora': str(c.monto_mora),
                    'vto': c.fecha_vencimiento.strftime('%d/%m/%Y') if c.fecha_vencimiento else '',
                    'estado': c.get_estado_display(),
                }
                for c in cuotas
            ]
        }

        return JsonResponse(data)

    elif request.method == 'POST':
        regularizacion_form = RegularizacionForm(request.POST, prefix='regularizacion_reg', instance=regularizacion)

        # Obtener o crear las reglas asociadas
        try:
            estructura_instance = ReglaEstructura.objects.get(regularizacion=regularizacion)
        except ReglaEstructura.DoesNotExist:
            estructura_instance = None

        try:
            mora_instance = ReglaMora.objects.get(regularizacion=regularizacion)
        except ReglaMora.DoesNotExist:
            mora_instance = None

        estructura_form = ReglaEstructuraRegularizacionForm(request.POST, prefix='estructura_reg', instance=estructura_instance)
        mora_form = ReglaMoraForm(request.POST, prefix='mora_reg', instance=mora_instance)

        if all([regularizacion_form.is_valid(), estructura_form.is_valid(), mora_form.is_valid()]):
            try:
                with transaction.atomic():
                    regularizacion = regularizacion_form.save()

                    estructura = estructura_form.save(commit=False)
                    estructura.regularizacion = regularizacion
                    estructura.save()

                    mora = mora_form.save(commit=False)
                    mora.regularizacion = regularizacion
                    mora.save()

                    # Actualizar las cuotas si fueron enviadas
                    cuotas_json = request.POST.get('cuotas_json')
                    if cuotas_json:
                        import json
                        from datetime import datetime
                        try:
                            # Eliminar cuotas existentes y crear las nuevas
                            CuotaRegularizacion.objects.filter(regularizacion=regularizacion).delete()

                            cuotas_data = json.loads(cuotas_json)
                            for cuota in cuotas_data:
                                fecha_vto = None
                                if cuota.get('vto'):
                                    try:
                                        fecha_vto = datetime.strptime(cuota['vto'], '%d/%m/%Y').date()
                                    except:
                                        pass

                                CuotaRegularizacion.objects.create(
                                    regularizacion=regularizacion,
                                    numero_cuota=cuota.get('nro', 0),
                                    fecha_vencimiento=fecha_vto,
                                    monto_base=cuota.get('base', 0),
                                    monto_interes=cuota.get('interes', 0),
                                    monto_cuota=cuota.get('monto_total', 0),
                                    monto_mora=cuota.get('mora', 0),
                                    estado='P'
                                )
                        except Exception as e:
                            print(f"Error al actualizar cuotas: {e}")

                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({'success': True, 'message': 'Regularización actualizada correctamente.'})

                messages.success(request, 'La regularización fue actualizada exitosamente.')
                return redirect('planes_list')

            except Exception as e:
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': str(e)}, status=500)
                messages.error(request, f'Ocurrió un error al actualizar: {e}')

        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': {
                        'regularizacion_reg': regularizacion_form.errors,
                        'estructura_reg': estructura_form.errors,
                        'mora_reg': mora_form.errors
                    }
                }, status=400)

            messages.error(request, 'Hay errores en los formularios.')

    return redirect('planes_list')


@login_required
def plan_editar(request, pk):
    """
    Editar un plan de pago existente con sus reglas de estructura y mora.
    Devuelve los datos del plan en formato JSON para cargar en el modal.
    Solo Administrador y Tesorero pueden hacer POST (editar).
    Consulta puede hacer GET (solo lectura).
    """
    # Verificar permisos para POST (edición)
    if request.method == 'POST':
        if not (request.user.is_superuser or request.user.groups.filter(name__in=['Administrador', 'Tesorero']).exists()):
            return JsonResponse({'ok': False, 'msg': 'No tiene permisos para editar planes'}, status=403)

    regularizacion = get_object_or_404(Regularizacion, pk=pk, tipo='Plan Normal')

    if request.method == 'GET':
        # Obtener las reglas asociadas
        try:
            estructura = ReglaEstructura.objects.get(regularizacion=regularizacion)
        except ReglaEstructura.DoesNotExist:
            estructura = None

        try:
            mora = ReglaMora.objects.get(regularizacion=regularizacion)
        except ReglaMora.DoesNotExist:
            mora = None

        # Obtener las cuotas existentes
        cuotas = CuotaRegularizacion.objects.filter(regularizacion=regularizacion).order_by('numero_cuota')

        # Preparar datos para enviar al frontend
        data = {
            'ok': True,
            'regularizacion': {
                'id': regularizacion.id,
                'nombre': regularizacion.nombre,
                'carrera': regularizacion.carrera,
                'modalidad': regularizacion.modalidad,
                'cohorte': regularizacion.cohorte,
            },
            'estructura': {
                'valor': str(estructura.valor) if estructura else '',
                'tasa': str(estructura.tasa) if estructura else '',
                'pago_inicial': str(estructura.pago_incial) if estructura else '',
                'cantidad_cuotas': estructura.cantidad_de_cuotas if estructura else '',
                'frecuencia': estructura.frecuencia_de_pago if estructura else '',
                'dia_vencimiento': estructura.dia_vencimiento if estructura else '',
            } if estructura else None,
            'mora': {
                'tipo_recargo': mora.tipo_de_recargo if mora else '',
                'cantidad': str(mora.cantidad_recargo) if mora else '',
                'frecuencia': mora.frecuencia_aplicacion if mora else '',
                'dias_gracia': mora.dias_gracia if mora else '',
                'veces_aplicacion': mora.veces_aplicacion if mora else '',
            } if mora else None,
            'cuotas': [
                {
                    'nro': c.numero_cuota,
                    'base': str(c.monto_base),
                    'interes': str(c.monto_interes),
                    'monto_total': str(c.monto_cuota),
                    'mora': str(c.monto_mora),
                    'vto': c.fecha_vencimiento.strftime('%d/%m/%Y') if c.fecha_vencimiento else '',
                    'estado': c.get_estado_display(),
                }
                for c in cuotas
            ]
        }

        return JsonResponse(data)

    elif request.method == 'POST':
        regularizacion_form = RegularizacionForm(request.POST, prefix='regularizacion', instance=regularizacion)

        # Obtener o crear las reglas asociadas
        try:
            estructura_instance = ReglaEstructura.objects.get(regularizacion=regularizacion)
        except ReglaEstructura.DoesNotExist:
            estructura_instance = None

        try:
            mora_instance = ReglaMora.objects.get(regularizacion=regularizacion)
        except ReglaMora.DoesNotExist:
            mora_instance = None

        estructura_form = ReglaEstructuraForm(request.POST, prefix='estructura', instance=estructura_instance)
        mora_form = ReglaMoraForm(request.POST, prefix='mora', instance=mora_instance)

        if all([regularizacion_form.is_valid(), estructura_form.is_valid(), mora_form.is_valid()]):
            try:
                with transaction.atomic():
                    regularizacion = regularizacion_form.save()

                    estructura = estructura_form.save(commit=False)
                    estructura.regularizacion = regularizacion
                    estructura.save()

                    mora = mora_form.save(commit=False)
                    mora.regularizacion = regularizacion
                    mora.save()

                    # Actualizar las cuotas si fueron enviadas
                    cuotas_json = request.POST.get('cuotas_json')
                    if cuotas_json:
                        import json
                        from datetime import datetime
                        try:
                            # Eliminar cuotas antiguas
                            CuotaRegularizacion.objects.filter(regularizacion=regularizacion).delete()

                            cuotas_data = json.loads(cuotas_json)
                            for cuota in cuotas_data:
                                # Parsear la fecha de vencimiento
                                fecha_vto = None
                                if cuota.get('vto'):
                                    try:
                                        fecha_vto = datetime.strptime(cuota['vto'], '%d/%m/%Y').date()
                                    except:
                                        try:
                                            fecha_vto = datetime.strptime(cuota['vto'], '%Y-%m-%d').date()
                                        except:
                                            pass

                                CuotaRegularizacion.objects.create(
                                    regularizacion=regularizacion,
                                    numero_cuota=cuota.get('nro', 0),
                                    fecha_vencimiento=fecha_vto,
                                    monto_base=cuota.get('base', 0),
                                    monto_interes=cuota.get('interes', 0),
                                    monto_cuota=cuota.get('monto_total', 0),
                                    monto_mora=cuota.get('mora', 0),
                                    estado='P'
                                )
                        except Exception as e:
                            print(f"Error al actualizar cuotas: {e}")

                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({'success': True, 'message': 'Plan actualizado correctamente.'})

                messages.success(request, 'El plan fue actualizado exitosamente.')
                return redirect('planes_list')

            except Exception as e:
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': str(e)}, status=500)
                messages.error(request, f'Ocurrió un error al actualizar: {e}')

        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': {
                        'regularizacion': regularizacion_form.errors,
                        'estructura': estructura_form.errors,
                        'mora': mora_form.errors
                    }
                }, status=400)

            messages.error(request, 'Hay errores en los formularios.')

    return redirect('planes_list')


@login_required
def regularizacion_clonar(request, pk):
    """Clonar una regularización o plan existente"""

    if request.method == "GET":
        # Intentar buscar en Regularizacion primero
        try:
            original = Regularizacion.objects.get(pk=pk)
            es_regularizacion = True
        except Regularizacion.DoesNotExist:
            # Si no está en Regularizacion, buscar en PlanPago
            try:
                original = PlanPago.objects.get(pk=pk)
                es_regularizacion = False
            except PlanPago.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Plan no encontrado'}, status=404)

        # Preparar datos básicos
        data = {
            'id': None,  # Nuevo registro
            'nombre': f"{original.nombre} (Copia)",
            'carrera': original.carrera,
            'cohorte': original.cohorte,
            'modalidad': original.modalidad,
            'estado': 'S',  # Suspendido por defecto
        }

        # Si es una regularización, agregar datos específicos
        if es_regularizacion:
            data['tipo'] = original.tipo

            # Agregar datos de estructura si existe
            if hasattr(original, 'estructura') and original.estructura:
                estructura = original.estructura
                data.update({
                    'pago_incial': str(estructura.pago_incial),
                    'cantidad_cuotas': estructura.cantidad_cuotas,
                    'interes': str(estructura.interes),
                    'vencimiento': estructura.vencimiento.strftime('%Y-%m-%d') if estructura.vencimiento else '',
                })

            # Agregar datos de mora si existe
            if hasattr(original, 'mora') and original.mora:
                mora = original.mora
                data.update({
                    'tipo_recargo': mora.tipo_recargo,
                    'cantidad_recargo': str(mora.cantidad_recargo),
                    'dias_gracia': mora.dias_gracia,
                })
        else:
            # Es un PlanPago antiguo
            data['tipo'] = original.tipo if hasattr(original, 'tipo') else 'P'

        return JsonResponse({'success': True, 'data': data})

    elif request.method == "POST":
        # Clonar directamente el plan
        try:
            # Intentar buscar en Regularizacion primero
            try:
                original = Regularizacion.objects.get(pk=pk)
                es_regularizacion = True
            except Regularizacion.DoesNotExist:
                # Si no está en Regularizacion, buscar en PlanPago
                try:
                    original = PlanPago.objects.get(pk=pk)
                    es_regularizacion = False
                except PlanPago.DoesNotExist:
                    return JsonResponse({'success': False, 'error': 'Plan no encontrado'}, status=404)

            if es_regularizacion:
                # Clonar regularización
                nueva = Regularizacion.objects.create(
                    nombre=f"{original.nombre} (Copia)",
                    carrera=original.carrera,
                    cohorte=original.cohorte,
                    modalidad=original.modalidad,
                    tipo=original.tipo,
                    estado='S'  # Suspendido por defecto
                )

                # Clonar estructura si existe
                if hasattr(original, 'reglas_estructura') and original.reglas_estructura.exists():
                    estructura_original = original.reglas_estructura.first()
                    ReglaEstructura.objects.create(
                        regularizacion=nueva,
                        origen_deuda=estructura_original.origen_deuda,
                        valor=estructura_original.valor,
                        tasa=estructura_original.tasa,
                        pago_incial=estructura_original.pago_incial,
                        cantidad_de_cuotas=estructura_original.cantidad_de_cuotas,
                        frecuencia_de_pago=estructura_original.frecuencia_de_pago,
                        dia_vencimiento=estructura_original.dia_vencimiento
                    )

                # Clonar mora si existe
                if hasattr(original, 'reglas_mora') and original.reglas_mora.exists():
                    mora_original = original.reglas_mora.first()
                    ReglaMora.objects.create(
                        regularizacion=nueva,
                        tipo_de_recargo=mora_original.tipo_de_recargo,
                        cantidad_recargo=mora_original.cantidad_recargo,
                        frecuencia_aplicacion=mora_original.frecuencia_aplicacion,
                        dias_gracia=mora_original.dias_gracia,
                        veces_aplicacion=mora_original.veces_aplicacion
                    )

                # Clonar cuotas si existen
                if hasattr(original, 'cuotas') and original.cuotas.exists():
                    for cuota_original in original.cuotas.all():
                        CuotaRegularizacion.objects.create(
                            regularizacion=nueva,
                            numero_cuota=cuota_original.numero_cuota,
                            fecha_vencimiento=cuota_original.fecha_vencimiento,
                            monto_base=cuota_original.monto_base,
                            monto_interes=cuota_original.monto_interes,
                            monto_cuota=cuota_original.monto_cuota,
                            monto_mora=cuota_original.monto_mora,
                            estado='P'  # Pendiente
                        )

                # Registrar en historial
                registrar_accion(
                    usuario=request.user,
                    accion="Clonó una regularización",
                    regularizacion=nueva,
                    descripcion=f"Clonó la regularización '{original.nombre}' → Nuevo: '{nueva.nombre}'"
                )

                return JsonResponse({
                    'success': True,
                    'message': f'Se clonó exitosamente: {nueva.nombre}',
                    'plan_id': nueva.id
                })
            else:
                # Clonar PlanPago antiguo
                nuevo = PlanPago.objects.create(
                    nombre=f"{original.nombre} (Copia)",
                    carrera=original.carrera,
                    cohorte=original.cohorte,
                    modalidad=original.modalidad,
                    tipo=original.tipo if hasattr(original, 'tipo') else 'P',
                    estado='S'
                )

                # Registrar en historial
                registrar_accion(
                    usuario=request.user,
                    accion="Clonó un plan",
                    plan=nuevo,
                    descripcion=f"Clonó el plan '{original.nombre}' → Nuevo: '{nuevo.nombre}'"
                )

                return JsonResponse({
                    'success': True,
                    'message': f'Se clonó exitosamente: {nuevo.nombre}',
                    'plan_id': nuevo.id
                })

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)


@login_required
def plan_clonar(request, pk):
    original = get_object_or_404(PlanPago, pk=pk)

    if request.method == "GET":
        # Si es petición AJAX, retornar JSON con los datos
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            data = {
                'id': None,  # Nuevo registro
                'tipo': original.tipo if hasattr(original, 'tipo') else 'P',
                'nombre': f"{original.nombre} (Copia)",
                'carrera': original.carrera,
                'cohorte': original.cohorte,
                'modalidad': original.modalidad,
                'estado': 'S',  # Suspendido por defecto
            }
            return JsonResponse({'success': True, 'data': data})

        # Si no es AJAX, usar el formulario antiguo
        data_inicial = {
            "nombre": f"{original.nombre} (Copia)",
            "carrera": original.carrera,
            "cohorte": original.cohorte,
            "modalidad": original.modalidad,
            "iEstado": True,
            "estado": "S",
        }
        form = PlanPagoForm(initial=data_inicial)
        return render(request, "planes/plan_form.html", {"form": form})

    elif request.method == "POST":
        form = PlanPagoForm(request.POST)
        if form.is_valid():
            nuevo_plan = form.save()   # ⚠️ Guardamos el nuevo plan y lo usamos abajo

            #  REGISTRAR ACCIÓN EN HISTORIAL
            registrar_accion(
                usuario=request.user,
                accion="Clonó un plan",
                plan=nuevo_plan,
                descripcion=f"Clonó el plan '{original}' → Nuevo: '{nuevo_plan}'"
            )

            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({"success": True})
            return redirect("planes_list")

    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)



# -------------------------------
# CRUD Cuotas
# -------------------------------
@login_required
def cuotas_list(request):
    # Obtener tanto PlanPago antiguos como Regularizaciones activas
    plans_antiguos = PlanPago.objects.filter(estado='A')
    regularizaciones = Regularizacion.objects.filter(estado='A')

    # Combinar ambas listas en un solo contexto
    # Nota: el template mostrará tanto planes antiguos como nuevos
    return render(request, "cuotas_list.html", {
        "plans": plans_antiguos,
        "regularizaciones": regularizaciones
    })


@login_required
def cuotas_data(request):
    plan_id = request.GET.get('plan')

    # Si hay un plan_id, verificar si es Regularizacion o PlanPago
    if plan_id:
        try:
            # Intentar buscar como Regularizacion primero
            regularizacion = Regularizacion.objects.get(pk=plan_id)
            # Es una regularización, buscar cuotas de CuotaRegularizacion
            data = [
                {
                    "id": c.id,
                    "plan": c.regularizacion.nombre,
                    "numero": c.numero_cuota,
                    "vencimiento": c.fecha_vencimiento.strftime("%Y-%m-%d") if c.fecha_vencimiento else "",
                    "monto": str(c.monto_cuota),
                }
                for c in CuotaRegularizacion.objects.filter(regularizacion_id=plan_id).order_by('numero_cuota')
            ]
        except Regularizacion.DoesNotExist:
            # No es regularización, intentar como PlanPago
            try:
                plan = PlanPago.objects.get(pk=plan_id)
                data = [
                    {
                        "id": c.id,
                        "plan": c.plan.nombre,
                        "numero": c.numero,
                        "vencimiento": c.vencimiento.strftime("%Y-%m-%d"),
                        "monto": str(c.monto),
                    }
                    for c in Cuota.objects.filter(plan_id=plan_id, iEstado=True).order_by('numero')
                ]
            except PlanPago.DoesNotExist:
                data = []
    else:
        # Sin filtro, mostrar todas las cuotas (comportamiento anterior)
        data = [
            {
                "id": c.id,
                "plan": c.plan.nombre,
                "numero": c.numero,
                "vencimiento": c.vencimiento.strftime("%Y-%m-%d"),
                "monto": str(c.monto),
            }
            for c in Cuota.objects.filter(iEstado=True)
        ]

    return JsonResponse({"data": data})


@require_POST
@login_required
def cuota_guardar(request):
    if not can_modify(request.user):
        return JsonResponse({"ok": False, "msg": "No tienes permisos para realizar esta acción"}, status=403)

    datos = request.POST
    cuota_id = datos.get("id")

    if cuota_id:
        try:
            cuota = Cuota.objects.get(pk=cuota_id)
            cuota.plan_id = datos.get("plan", cuota.plan_id)
            cuota.numero = datos.get("numero", cuota.numero)
            cuota.vencimiento = datos.get("vencimiento", cuota.vencimiento)
            cuota.monto = datos.get("monto", cuota.monto)
            cuota.save()
            return JsonResponse({"ok": True, "msg": "Cuota actualizada"})
        except Cuota.DoesNotExist:
            return JsonResponse({"ok": False, "msg": "Cuota no encontrada"}, status=404)
    else:
        Cuota.objects.create(
            plan_id=datos.get("plan"),
            numero=datos.get("numero"),
            vencimiento=datos.get("vencimiento"),
            monto=datos.get("monto"),
        )
        return JsonResponse({"ok": True, "msg": "Cuota creada"})


@require_POST
@login_required
def cuota_borrar(request, pk):
    if not can_delete(request.user):
        return JsonResponse({"ok": False, "msg": "No tienes permisos para eliminar cuotas"}, status=403)

    try:
        cuota = Cuota.objects.get(pk=pk)
        cuota.iEstado = False
        cuota.save()
        return JsonResponse({"ok": True, "msg": "Cuota eliminada"})
    except Cuota.DoesNotExist:
        return JsonResponse({"ok": False, "msg": "Cuota no encontrada"}, status=404)


# -------------------------------
# EXPORTAR PLANES
# -------------------------------
@login_required
def exportar_planes_pdf(request):
    pdf = PDF()
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
    usuario = request.user.username
    pdf.cell(0, 10, f"Listado de Planes de Pago", ln=True, align="C")
    pdf.set_font("Arial", "I", 10)
    pdf.cell(0, 10, f"Generado por: {usuario} - Fecha: {fecha}", ln=True, align="C")
    pdf.ln(10)

    # Headers con anchos ajustados para que quepan en la página
    pdf.set_font("Arial", "B", 9)
    pdf.cell(30, 10, "Tipo", 1, 0, "C")
    pdf.cell(50, 10, "Nombre", 1, 0, "C")
    pdf.cell(40, 10, "Carrera", 1, 0, "C")
    pdf.cell(30, 10, "Cohorte", 1, 0, "C")
    pdf.cell(30, 10, "Modalidad", 1, 1, "C")

    pdf.set_font("Arial", size=8)

    # Exportar Regularizaciones (nuevos planes)
    for reg in Regularizacion.objects.filter(estado='A'):
        tipo = reg.get_tipo_display() if hasattr(reg, 'get_tipo_display') else reg.tipo

        # Acortar strings si son muy largos
        tipo_str = str(tipo)[:15] if len(str(tipo)) > 15 else str(tipo)
        nombre = str(reg.nombre)[:30] if len(str(reg.nombre)) > 30 else str(reg.nombre)
        carrera = str(reg.carrera)[:25] if len(str(reg.carrera)) > 25 else str(reg.carrera)
        cohorte = str(reg.cohorte)[:15] if len(str(reg.cohorte)) > 15 else str(reg.cohorte)
        modalidad = str(reg.modalidad)[:15] if len(str(reg.modalidad)) > 15 else str(reg.modalidad)

        pdf.cell(30, 8, tipo_str, 1)
        pdf.cell(50, 8, nombre, 1)
        pdf.cell(40, 8, carrera, 1)
        pdf.cell(30, 8, cohorte, 1)
        pdf.cell(30, 8, modalidad, 1)
        pdf.ln()

    # Exportar PlanPago antiguos (si existen)
    for p in PlanPago.objects.filter(estado='A'):
        nombre = str(p.nombre)[:30] if len(str(p.nombre)) > 30 else str(p.nombre)
        carrera = str(p.carrera)[:25] if len(str(p.carrera)) > 25 else str(p.carrera)
        cohorte = str(p.cohorte)[:15] if len(str(p.cohorte)) > 15 else str(p.cohorte)
        modalidad = str(p.modalidad)[:15] if len(str(p.modalidad)) > 15 else str(p.modalidad)

        pdf.cell(30, 8, "Plan Antiguo", 1)
        pdf.cell(50, 8, nombre, 1)
        pdf.cell(40, 8, carrera, 1)
        pdf.cell(30, 8, cohorte, 1)
        pdf.cell(30, 8, modalidad, 1)
        pdf.ln()

    pdf_bytes = pdf.output(dest="S")
    response = HttpResponse(bytes(pdf_bytes), content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="planes_pago.pdf"'
    return response


@login_required
def exportar_planes_csv(request):
    response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
    response["Content-Disposition"] = 'attachment; filename="planes_pago.csv"'

    writer = csv.writer(response, delimiter=";")
    writer.writerow(["Tipo de Plan", "Nombre", "Carrera", "Cohorte", "Modalidad"])

    # Exportar Regularizaciones (nuevos planes)
    for reg in Regularizacion.objects.filter(estado='A'):
        tipo = reg.get_tipo_display() if hasattr(reg, 'get_tipo_display') else reg.tipo
        writer.writerow([tipo, reg.nombre, reg.carrera, reg.cohorte, reg.modalidad])

    # Exportar PlanPago antiguos (si existen)
    for p in PlanPago.objects.filter(estado='A'):
        writer.writerow(["Plan Antiguo", p.nombre, p.carrera, p.cohorte, p.modalidad])

    return response


@login_required
def exportar_planes_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planes de Pago"

    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    bold_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["Tipo de Plan", "Nombre", "Carrera", "Cohorte", "Modalidad"]
    ws.append(headers)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Exportar Regularizaciones (nuevos planes)
    for reg in Regularizacion.objects.filter(estado='A'):
        tipo = reg.get_tipo_display() if hasattr(reg, 'get_tipo_display') else reg.tipo
        ws.append([tipo, reg.nombre, reg.carrera, reg.cohorte, reg.modalidad])

    # Exportar PlanPago antiguos (si existen)
    for p in PlanPago.objects.filter(estado='A'):
        ws.append(["Plan Antiguo", p.nombre, p.carrera, p.cohorte, p.modalidad])

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="planes_pago.xlsx"'
    wb.save(response)
    return response


@login_required
def imprimir_planes(request):
    # Preparar datos de planes con toda la información necesaria
    planes_data = []

    # Agregar Regularizaciones (nuevos planes)
    for reg in Regularizacion.objects.filter(estado='A'):
        tipo = reg.get_tipo_display() if hasattr(reg, 'get_tipo_display') else reg.tipo
        planes_data.append({
            'tipo': tipo,
            'nombre': reg.nombre,
            'carrera': reg.carrera,
            'cohorte': reg.cohorte,
            'modalidad': reg.modalidad
        })

    # Agregar PlanPago antiguos (si existen)
    for p in PlanPago.objects.filter(estado='A'):
        planes_data.append({
            'tipo': "Plan Antiguo",
            'nombre': p.nombre,
            'carrera': p.carrera,
            'cohorte': p.cohorte,
            'modalidad': p.modalidad
        })

    return render(request, "imprimir_planes.html", {"planes": planes_data})


# -------------------------------
# EXPORTAR CUOTAS
# -------------------------------
@login_required
def exportar_cuotas_pdf(request):
    pdf = PDF()
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
    usuario = request.user.username
    pdf.cell(0, 10, "Listado de Cuotas", ln=True, align="C")
    pdf.set_font("Arial", "I", 10)
    pdf.cell(0, 10, f"Generado por: {usuario} - Fecha: {fecha}", ln=True, align="C")
    pdf.ln(10)

    pdf.set_font("Arial", "B", 10)
    pdf.cell(30, 10, "Plan", 1, 0, "C")
    pdf.cell(30, 10, "Número", 1, 0, "C")
    pdf.cell(40, 10, "Vencimiento", 1, 0, "C")
    pdf.cell(40, 10, "Monto", 1, 1, "C")

    pdf.set_font("Arial", size=10)
    for c in Cuota.objects.filter(iEstado=True):
        pdf.cell(30, 10, str(c.plan.nombre), 1)
        pdf.cell(30, 10, str(c.numero), 1)
        pdf.cell(40, 10, c.vencimiento.strftime("%Y-%m-%d"), 1)
        pdf.cell(40, 10, str(c.monto), 1)
        pdf.ln()

    pdf_bytes = pdf.output(dest="S")
    response = HttpResponse(bytes(pdf_bytes), content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="cuotas.pdf"'
    return response


@login_required
def exportar_cuotas_csv(request):
    response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
    response["Content-Disposition"] = 'attachment; filename="cuotas.csv"'

    writer = csv.writer(response, delimiter=";")
    writer.writerow(["Plan", "Número de Cuota", "Fecha de Vencimiento", "Monto ($)"])

    for c in Cuota.objects.filter(iEstado=True):
        writer.writerow([c.plan.nombre, c.numero, c.vencimiento.strftime("%d/%m/%Y"), c.monto])

    return response


@login_required
def exportar_cuotas_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cuotas"

    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    bold_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["Plan", "Número", "Vencimiento", "Monto"]
    ws.append(headers)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for c in Cuota.objects.filter(iEstado=True):
        ws.append([c.plan.nombre, c.numero, c.vencimiento.strftime("%d/%m/%Y"), float(c.monto)])

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="cuotas.xlsx"'
    wb.save(response)
    return response


@login_required
def imprimir_cuotas(request):
    cuotas = Cuota.objects.filter(iEstado=True)
    return render(request, "imprimir_cuotas.html", {"cuotas": cuotas})


# -------------------------------
# Planes suspendidos
# -------------------------------
@login_required
def planes_suspendidos(request):

    # --- Planes ---
    suspendidos_planes = PlanPago.objects.filter(estado='S')
    desactivados_planes = PlanPago.objects.filter(estado='D')

    # --- Regularizaciones ---
    suspendidos_regularizaciones = Regularizacion.objects.filter(estado='S')
    desactivados_regularizaciones = Regularizacion.objects.filter(estado='D')

    # --- Marcar cada objeto con su tipo ---
    for p in suspendidos_planes:
        p.tipo_backend = "plan"
    for p in desactivados_planes:
        p.tipo_backend = "plan"

    for r in suspendidos_regularizaciones:
        r.tipo_backend = "regularizacion"
    for r in desactivados_regularizaciones:
        r.tipo_backend = "regularizacion"

    # --- Combinar ---
    suspendidos = list(suspendidos_planes) + list(suspendidos_regularizaciones)
    desactivados = list(desactivados_planes) + list(desactivados_regularizaciones)

    return render(request, "planes_suspendidos.html", {
        "suspendidos": suspendidos,
        "desactivados": desactivados
    })
@login_required
def plan_ver(request, pk):
    plan = get_object_or_404(PlanPago, pk=pk)
    return render(request, "planes/ver_plan.html", {"plan": plan})


@login_required
def plan_ver_detalle(request, pk):
    """
    Vista para ver los detalles de un plan o regularización en formato JSON
    """
    tipo = request.GET.get('tipo', 'plan')

    try:
        if tipo == 'regularizacion':
            obj = Regularizacion.objects.get(pk=pk)
            plan_data = {
                'id': obj.id,
                'nombre': obj.nombre,
                'carrera': obj.carrera,
                'cohorte': obj.cohorte,
                'modalidad': obj.modalidad,
                'estado': obj.estado,
                'tipo': obj.tipo,  # Agregamos el tipo para identificarlo en el frontend
                'fecha_creacion': obj.fecha_creacion.strftime('%d/%m/%Y %H:%M') if hasattr(obj, 'fecha_creacion') and obj.fecha_creacion else 'N/A',
            }

            # Buscar estructura asociada
            try:
                estructura = ReglaEstructura.objects.filter(regularizacion=obj).first()
                if estructura:
                    plan_data['estructura'] = {
                        'origen_deuda': estructura.get_origen_deuda_display() if estructura.origen_deuda else 'N/A',
                        'valor': str(estructura.valor),
                        'tasa': str(estructura.tasa),
                        'pago_inicial': str(estructura.pago_incial),
                        'cantidad_cuotas': estructura.cantidad_de_cuotas,
                        'frecuencia': estructura.get_frecuencia_de_pago_display(),
                        'dia_vencimiento': estructura.dia_vencimiento,
                    }
            except:
                pass

            # Buscar reglas de mora asociadas
            try:
                mora = ReglaMora.objects.filter(regularizacion=obj).first()
                if mora:
                    plan_data['mora'] = {
                        'tipo': mora.get_tipo_de_recargo_display() if mora.tipo_de_recargo else 'N/A',
                        'cantidad': str(mora.cantidad_recargo) if mora.cantidad_recargo else 'N/A',
                        'frecuencia': mora.get_frecuencia_aplicacion_display() if mora.frecuencia_aplicacion else 'N/A',
                        'dias_gracia': mora.dias_gracia if mora.dias_gracia else 0,
                        'veces_aplicacion': mora.veces_aplicacion if mora.veces_aplicacion else 0,
                    }
            except:
                pass

            # Buscar cuotas asociadas
            try:
                from .models import CuotaRegularizacion
                cuotas = CuotaRegularizacion.objects.filter(regularizacion=obj).order_by('numero_cuota')
                if cuotas.exists():
                    plan_data['cuotas'] = [
                        {
                            'numero': c.numero_cuota,
                            'fecha_vencimiento': c.fecha_vencimiento.strftime('%d/%m/%Y') if c.fecha_vencimiento else 'N/A',
                            'base': str(c.monto_base),
                            'interes': str(c.monto_interes),
                            'monto': str(c.monto_cuota),
                            'mora': str(c.monto_mora),
                            'estado': c.get_estado_display(),
                        }
                        for c in cuotas
                    ]
            except Exception as e:
                print(f"Error al obtener cuotas: {e}")

        else:  # plan
            # Para planes normales, buscar en la tabla Regularizacion con tipo='Plan Normal'
            obj = Regularizacion.objects.get(pk=pk, tipo='Plan Normal')
            plan_data = {
                'id': obj.id,
                'nombre': obj.nombre,
                'carrera': obj.carrera,
                'cohorte': obj.cohorte,
                'modalidad': obj.modalidad,
                'estado': obj.estado,
                'tipo': 'Plan Normal',  # Agregamos el tipo para identificarlo en el frontend
                'fecha_creacion': obj.fecha_creacion.strftime('%d/%m/%Y %H:%M') if hasattr(obj, 'fecha_creacion') and obj.fecha_creacion else 'N/A',
            }

            # Buscar estructura asociada (NO tiene origen_deuda)
            try:
                estructura = ReglaEstructura.objects.filter(regularizacion=obj).first()
                if estructura:
                    plan_data['estructura'] = {
                        # NO incluir origen_deuda para planes normales
                        'valor': str(estructura.valor),
                        'tasa': str(estructura.tasa),
                        'pago_inicial': str(estructura.pago_incial),
                        'cantidad_cuotas': estructura.cantidad_de_cuotas,
                        'frecuencia': estructura.get_frecuencia_de_pago_display(),
                        'dia_vencimiento': estructura.dia_vencimiento,
                    }
            except:
                pass

            # Buscar reglas de mora asociadas
            try:
                mora = ReglaMora.objects.filter(regularizacion=obj).first()
                if mora:
                    plan_data['mora'] = {
                        'tipo': mora.get_tipo_de_recargo_display() if mora.tipo_de_recargo else 'N/A',
                        'cantidad': str(mora.cantidad_recargo) if mora.cantidad_recargo else 'N/A',
                        'frecuencia': mora.get_frecuencia_aplicacion_display() if mora.frecuencia_aplicacion else 'N/A',
                        'dias_gracia': mora.dias_gracia if mora.dias_gracia else 0,
                        'veces_aplicacion': mora.veces_aplicacion if mora.veces_aplicacion else 0,
                    }
            except:
                pass

            # Buscar cuotas asociadas
            try:
                cuotas = CuotaRegularizacion.objects.filter(regularizacion=obj).order_by('numero_cuota')
                if cuotas.exists():
                    plan_data['cuotas'] = [
                        {
                            'numero': c.numero_cuota,
                            'fecha_vencimiento': c.fecha_vencimiento.strftime('%d/%m/%Y') if c.fecha_vencimiento else 'N/A',
                            'base': str(c.monto_base),
                            'interes': str(c.monto_interes),
                            'monto': str(c.monto_cuota),
                            'mora': str(c.monto_mora),
                            'estado': c.get_estado_display(),
                        }
                        for c in cuotas
                    ]
            except Exception as e:
                print(f"Error al obtener cuotas: {e}")

        return JsonResponse({'ok': True, 'plan': plan_data})

    except (PlanPago.DoesNotExist, Regularizacion.DoesNotExist):
        return JsonResponse({'ok': False, 'msg': f'{tipo.capitalize()} no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'ok': False, 'msg': f'Error al obtener detalles: {str(e)}'}, status=500)


@require_POST
@login_required
def plan_suspendido(request, pk):
    """
    Acción que marca iEstado=False (uso legacy). Dejéla si alguna parte del sistema la llama.
    """
    try:
        plan = PlanPago.objects.get(pk=pk)
        plan.iEstado = False
        plan.save()
        return JsonResponse({"ok": True, "msg": "Plan suspendido correctamente"})
    except PlanPago.DoesNotExist:
        return JsonResponse({"ok": False, "msg": "Plan no encontrado"}, status=404)


# -------------------------------
# DESACTIVAR / REACTIVAR PLANES
# -------------------------------

@require_POST
@login_required
def plan_suspender(request, pk):
    if not can_delete(request.user):
        return JsonResponse({"ok": False, "msg": "No tienes permisos para suspender planes"}, status=403)

    # 1. Intentar suspender un PlanPago
    try:
        plan = PlanPago.objects.get(pk=pk)
        plan.estado = 'S'
        plan.iEstado = False
        plan.save()

        #  Registrar historial
        registrar_accion(
            usuario=request.user,
            accion="Suspendió un plan",
            plan=plan,
            descripcion=f"El plan '{plan}' fue suspendido."
        )

        return JsonResponse({"ok": True, "msg": "Plan suspendido correctamente"})
    except PlanPago.DoesNotExist:
        pass

    # 2. Intentar suspender una Regularización
    try:
        reg = Regularizacion.objects.get(pk=pk)
        reg.estado = 'S'
        reg.save()

        #  Registrar historial
        registrar_accion(
            usuario=request.user,
            accion="Suspendió una regularización",
            plan=None,
            descripcion=f"Regularización '{reg}' fue suspendida."
        )

        return JsonResponse({"ok": True, "msg": "Regularización suspendida correctamente"})
    except Regularizacion.DoesNotExist:
        pass

    # 3. Si no existe en ninguna tabla
    return JsonResponse({"ok": False, "msg": "Objeto no encontrado"}, status=404)


@require_POST
@login_required
def desactivar_objeto(request):
    tipo = request.POST.get("tipo")
    pk = request.POST.get("id")

    print("DEBUG tipo recibido:", repr(tipo))
    print("DEBUG id recibido:", repr(pk))

    if tipo not in ["plan", "regularizacion"]:
        return JsonResponse({"ok": False, "msg": f"Tipo inválido: {tipo}"}, status=400)

    Model = PlanPago if tipo == "plan" else Regularizacion

    try:
        obj = Model.objects.get(pk=pk)
        obj.estado = "D"   # D = desactivado
        obj.save()

        # -------------------------------------------
        # REGISTRAR EN HISTORIAL
        # -------------------------------------------
        registrar_accion(
            usuario=request.user,
            accion=f"Desactivó un {tipo}",
            plan=obj if tipo == "plan" else None,
            descripcion=f"{tipo.capitalize()} desactivado: {obj}"
        )
        # -------------------------------------------

        return JsonResponse({"ok": True, "msg": f"{tipo.capitalize()} desactivado correctamente"})
    
    except Model.DoesNotExist:
        return JsonResponse({"ok": False, "msg": "Objeto no encontrado"}, status=404)

    except Exception as e:
        print("ERROR inesperado:", e)
        return JsonResponse({"ok": False, "msg": f"Error inesperado: {str(e)}"}, status=500)


# @require_POST -------ESTE ES EL REACTIVAR VIEJO, LO DEJO COMENTADO POR LAS DUDAS-------
# @login_required
# def plan_reactivar(request, pk):
#     """
#     Reactiva un plan desactivado o suspendido (estado='A' y iEstado=True).
#     """
#     if not can_delete(request.user):
#         return JsonResponse({"ok": False, "msg": "No tienes permisos para reactivar planes"}, status=403)

#     try:
#         plan = PlanPago.objects.get(pk=pk)
#         plan.estado = 'A'
#         plan.iEstado = True
#         plan.save()
#         return JsonResponse({"ok": True, "msg": "Plan reactivado correctamente"})
#     except PlanPago.DoesNotExist:
#         return JsonResponse({"ok": False, "msg": "Plan no encontrado"}, status=404)
@require_POST
@login_required
def reactivar_objeto(request):
    tipo = request.POST.get("tipo")
    pk = request.POST.get("id")

    if tipo not in ["plan", "regularizacion"]:
        return JsonResponse({"ok": False, "msg": "Tipo inválido"}, status=400)

    Model = PlanPago if tipo == "plan" else Regularizacion

    try:
        obj = Model.objects.get(pk=pk)
        obj.estado = "A"
        obj.save()

        # -------------------------------------------
        # REGISTRAR EN HISTORIAL
        # -------------------------------------------
        registrar_accion(
            usuario=request.user,
            accion=f"Reactivó un {tipo}",
            plan=obj if tipo == "plan" else None,
            descripcion=f"{tipo.capitalize()} reactivado: {obj}"
        )
        # -------------------------------------------

        return JsonResponse({"ok": True, "msg": f"{tipo.capitalize()} reactivado correctamente"})

    except Model.DoesNotExist:
        return JsonResponse({"ok": False, "msg": "Objeto no encontrado"}, status=404)


@login_required
@group_required('Administrador', 'Tesorero')
def historial(request):
    return render(request, "historial.html")

@login_required
@group_required('Administrador', 'Tesorero')
def historial_ajax(request):
    historial = HistorialAccion.objects.select_related("usuario", "plan").order_by("-fecha")

    data = []
    for h in historial:
        fecha_local = timezone.localtime(h.fecha)  # 👈 acá convertimos a tu zona

        # Mostrar nombre y apellido en lugar de username
        nombre_completo = f"{h.usuario.first_name} {h.usuario.last_name}" if h.usuario else "—"

        data.append({
            "fecha": fecha_local.strftime("%d/%m/%Y %H:%M"),
            "usuario": nombre_completo,
            "accion": h.accion,
            "plan": str(h.plan) if h.plan else "—",
            "descripcion": h.descripcion,
        })

    return JsonResponse({"data": data})



# -------------------------------
# Gestión de Usuarios y Roles
# -------------------------------
@login_required
@group_required('Administrador')
def usuarios_list(request):
    """Vista principal de gestión de usuarios (solo para administradores)"""
    from django.contrib.auth.models import User, Group

    return render(request, "usuarios_list.html")


@login_required
@group_required('Administrador')
def usuarios_data(request):
    """Endpoint JSON para datatables de usuarios"""
    from django.contrib.auth.models import User

    usuarios = User.objects.all().prefetch_related('groups')

    data = []
    for user in usuarios:
        grupos_list = list(user.groups.all())
        grupos_nombres = ", ".join([g.name for g in grupos_list]) or "Sin rol"
        grupo_id = grupos_list[0].id if grupos_list else None

        data.append({
            "id": user.id,
            "username": user.username,
            "email": user.email,
            "first_name": user.first_name,
            "last_name": user.last_name,
            "is_active": user.is_active,
            "grupos": grupos_nombres,
            "grupo_id": grupo_id,
        })

    return JsonResponse({"data": data})


@login_required
@group_required('Administrador')
def grupos_data(request):
    """Endpoint JSON para obtener grupos/roles disponibles"""
    from django.contrib.auth.models import Group

    print("=== DEBUG grupos_data ===")
    print(f"Usuario: {request.user.username}")
    print(f"Es superuser: {request.user.is_superuser}")
    print(f"Grupos del usuario: {list(request.user.groups.values_list('name', flat=True))}")

    grupos = Group.objects.all().values('id', 'name')
    grupos_list = list(grupos)

    print(f"Grupos disponibles en DB: {grupos_list}")
    print(f"Cantidad de grupos: {len(grupos_list)}")

    return JsonResponse({"grupos": grupos_list})


@require_POST
@login_required
@group_required('Administrador')
def usuario_crear(request):
    """Crear nuevo usuario con rol asignado"""
    from django.contrib.auth.models import User, Group

    try:
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        password = request.POST.get('password', '').strip()
        password_confirm = request.POST.get('password_confirm', '').strip()
        grupo_id = request.POST.get('grupo')

        # Validaciones
        if not username:
            return JsonResponse({"ok": False, "msg": "El nombre de usuario es obligatorio"}, status=400)

        if User.objects.filter(username=username).exists():
            return JsonResponse({"ok": False, "msg": "El nombre de usuario ya existe"}, status=400)

        if email and User.objects.filter(email=email).exists():
            return JsonResponse({"ok": False, "msg": "El email ya está registrado"}, status=400)

        if not password:
            return JsonResponse({"ok": False, "msg": "La contraseña es obligatoria"}, status=400)

        if password != password_confirm:
            return JsonResponse({"ok": False, "msg": "Las contraseñas no coinciden"}, status=400)

        if len(password) < 6:
            return JsonResponse({"ok": False, "msg": "La contraseña debe tener al menos 6 caracteres"}, status=400)

        # Crear usuario
        with transaction.atomic():
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name
            )

            # Asignar grupo/rol si fue seleccionado
            if grupo_id:
                try:
                    grupo = Group.objects.get(id=grupo_id)
                    user.groups.add(grupo)
                except Group.DoesNotExist:
                    pass

        return JsonResponse({
            "ok": True,
            "msg": f"Usuario '{username}' creado correctamente"
        })

    except Exception as e:
        return JsonResponse({
            "ok": False,
            "msg": f"Error al crear usuario: {str(e)}"
        }, status=500)


@require_POST
@login_required
@group_required('Administrador')
def usuario_editar(request, pk):
    """Editar usuario existente"""
    from django.contrib.auth.models import User, Group

    try:
        user = get_object_or_404(User, pk=pk)

        email = request.POST.get('email', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        is_active = request.POST.get('is_active') == 'true'
        grupo_id = request.POST.get('grupo')

        # Validar email único (excepto el actual)
        if email and User.objects.filter(email=email).exclude(pk=pk).exists():
            return JsonResponse({"ok": False, "msg": "El email ya está registrado"}, status=400)

        # Actualizar datos
        with transaction.atomic():
            user.email = email
            user.first_name = first_name
            user.last_name = last_name
            user.is_active = is_active
            user.save()

            # Actualizar grupo
            user.groups.clear()
            if grupo_id:
                try:
                    grupo = Group.objects.get(id=grupo_id)
                    user.groups.add(grupo)
                except Group.DoesNotExist:
                    pass

        return JsonResponse({
            "ok": True,
            "msg": f"Usuario '{user.username}' actualizado correctamente"
        })

    except Exception as e:
        return JsonResponse({
            "ok": False,
            "msg": f"Error al actualizar usuario: {str(e)}"
        }, status=500)


@require_POST
@login_required
@group_required('Administrador')
def usuario_eliminar(request, pk):
    """Desactivar usuario (no se elimina físicamente)"""
    from django.contrib.auth.models import User

    try:
        user = get_object_or_404(User, pk=pk)

        # No permitir desactivar el propio usuario
        if user.id == request.user.id:
            return JsonResponse({
                "ok": False,
                "msg": "No puedes desactivar tu propio usuario"
            }, status=400)

        user.is_active = False
        user.save()

        return JsonResponse({
            "ok": True,
            "msg": f"Usuario '{user.username}' desactivado correctamente"
        })

    except Exception as e:
        return JsonResponse({
            "ok": False,
            "msg": f"Error al desactivar usuario: {str(e)}"
        }, status=500)
